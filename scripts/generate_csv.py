"""
generate_csv.py - Final CSV/Excel output generation

Generates human-readable CSV and Excel reports from cross-reference results.
Excel output includes clickable hyperlinks in the career history columns.
"""

import csv
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from cross_reference import (
    cross_reference_all_coaches,
    format_overlaps_summary,
    load_json_file
)
from logger import setup_logger, get_logger, log_section, log_summary, get_log_file_path

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def format_career_history(career_history: List[Dict], year_start: int = 2020) -> str:
    """
    Format career history into a readable string.

    Args:
        career_history: List of career stint dictionaries
        year_start: Earliest year to include

    Returns:
        Formatted string like "Colorado (2024-present), Broncos (2022-2023)"
    """
    if not career_history:
        return "No career history found"

    parts = []
    for stint in career_history:
        school = stint.get("school", "Unknown")
        years = stint.get("years", "Unknown")

        # Try to extract start year to filter
        try:
            start_year = int(years.split("-")[0])
            if start_year < year_start:
                # Check end year
                end_part = years.split("-")[1]
                if end_part.lower() == "present":
                    pass  # Include it
                else:
                    end_year = int(end_part)
                    if end_year < year_start:
                        continue  # Skip entirely
        except (ValueError, IndexError):
            pass  # Include if we can't parse

        parts.append(f"{school} ({years})")

    return ", ".join(parts) if parts else "No recent career history"


def format_source_urls(career_history: List[Dict]) -> str:
    """
    Extract and format source URLs from career history.

    Args:
        career_history: List of career stint dictionaries

    Returns:
        Pipe-separated list of unique source URLs
    """
    urls = set()
    for stint in career_history:
        url = stint.get("source_url")
        if url:
            urls.add(url)

    return " | ".join(sorted(urls)) if urls else ""


def determine_data_quality(career_history: List[Dict], research_status: str) -> str:
    """
    Determine overall data quality for a coach.

    Args:
        career_history: List of career stint dictionaries
        research_status: Research status string

    Returns:
        "VERIFIED", "PARTIAL", or "UNVERIFIED"
    """
    if research_status in ["NOT_FOUND", "AMBIGUOUS"]:
        return "UNVERIFIED"

    if not career_history:
        return "UNVERIFIED"

    # Count entries with source URLs
    with_source = sum(1 for stint in career_history if stint.get("source_url"))
    total = len(career_history)

    if with_source == total:
        return "VERIFIED"
    elif with_source > 0:
        return "PARTIAL"
    else:
        return "UNVERIFIED"


def get_career_entries_with_urls(career_history: List[Dict], year_start: int = 2020) -> List[Tuple[str, str]]:
    """
    Extract career entries with their source URLs for Excel hyperlinks.

    Args:
        career_history: List of career stint dictionaries
        year_start: Earliest year to include

    Returns:
        List of tuples: (display_text, source_url)
    """
    if not career_history:
        return []

    entries = []
    for stint in career_history:
        school = stint.get("school", "Unknown")
        years = stint.get("years", "Unknown")
        source_url = stint.get("source_url", "")

        # Try to extract start year to filter
        try:
            start_year = int(years.split("-")[0])
            if start_year < year_start:
                end_part = years.split("-")[1]
                if end_part.lower() == "present":
                    pass  # Include it
                else:
                    end_year = int(end_part)
                    if end_year < year_start:
                        continue  # Skip entirely
        except (ValueError, IndexError):
            pass  # Include if we can't parse

        display_text = f"{school} ({years})"
        entries.append((display_text, source_url))

    return entries


def generate_excel_report(
    results: List[Dict],
    output_path: str,
    config: Dict = None
) -> str:
    """
    Generate an Excel report with clickable hyperlinks in career history.

    Args:
        results: List of cross-reference result dictionaries
        output_path: Path to write Excel file
        config: Configuration dictionary

    Returns:
        Path to generated Excel file
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")

    config = config or {}
    year_start = config.get("year_range", {}).get("start", 2020)

    # Determine max career entries across all coaches
    max_career_entries = 0
    for result in results:
        entries = get_career_entries_with_urls(result.get("career_history", []), year_start)
        max_career_entries = max(max_career_entries, len(entries))

    # Ensure at least 3 career columns
    max_career_entries = max(max_career_entries, 3)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Coach NMDP Cross-Reference"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    overlap_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Build headers
    headers = ["Coach Name", "Current School", "Current Position"]
    for i in range(max_career_entries):
        headers.append(f"Career {i+1}")
    headers.extend(["NMDP Overlap", "Overlap Details", "Data Quality"])

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, result in enumerate(results, 2):
        career_history = result.get("career_history", [])
        career_entries = get_career_entries_with_urls(career_history, year_start)
        has_overlap = result.get("has_overlap", False)

        # Basic info columns
        ws.cell(row=row_idx, column=1, value=result.get("coach_name", "Unknown"))
        ws.cell(row=row_idx, column=2, value=result.get("current_school", "Unknown"))
        ws.cell(row=row_idx, column=3, value=result.get("current_position", "Unknown"))

        # Career columns with hyperlinks
        for i in range(max_career_entries):
            col_idx = 4 + i
            if i < len(career_entries):
                display_text, source_url = career_entries[i]
                cell = ws.cell(row=row_idx, column=col_idx, value=display_text)
                if source_url:
                    cell.hyperlink = source_url
                    cell.font = link_font
            else:
                ws.cell(row=row_idx, column=col_idx, value="")

        # Overlap columns
        overlap_col = 4 + max_career_entries
        ws.cell(row=row_idx, column=overlap_col, value="YES" if has_overlap else "NO")
        ws.cell(row=row_idx, column=overlap_col + 1, value=format_overlaps_summary(result.get("overlaps", [])))
        ws.cell(row=row_idx, column=overlap_col + 2, value=determine_data_quality(career_history, result.get("research_status", "")))

        # Apply borders and highlighting to all cells in row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if has_overlap:
                cell.fill = overlap_fill

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(results) + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save workbook
    wb.save(output_path)

    return output_path


def generate_csv_report(
    results: List[Dict],
    output_path: str,
    config: Dict = None
) -> str:
    """
    Generate a CSV report from cross-reference results.

    Args:
        results: List of cross-reference result dictionaries
        output_path: Path to write CSV file
        config: Configuration dictionary

    Returns:
        Path to generated CSV file
    """
    config = config or {}
    year_start = config.get("year_range", {}).get("start", 2020)

    # Define CSV columns
    fieldnames = [
        "Coach Name",
        "Current School",
        "Current Position",
        f"Career History ({year_start}-Present)",
        "Source URLs",
        "NMDP Overlap",
        "Overlap Details",
        "Data Quality"
    ]

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for result in results:
            career_history = result.get("career_history", [])

            row = {
                "Coach Name": result.get("coach_name", "Unknown"),
                "Current School": result.get("current_school", "Unknown"),
                "Current Position": result.get("current_position", "Unknown"),
                f"Career History ({year_start}-Present)": format_career_history(career_history, year_start),
                "Source URLs": format_source_urls(career_history),
                "NMDP Overlap": "YES" if result.get("has_overlap") else "NO",
                "Overlap Details": format_overlaps_summary(result.get("overlaps", [])),
                "Data Quality": determine_data_quality(career_history, result.get("research_status", ""))
            }

            writer.writerow(row)

    return output_path


def generate_report_for_school(
    school_name: str,
    cache_base_dir: str,
    output_base_dir: str,
    nmdp_db_path: str,
    aliases_path: str,
    config_path: str,
    enable_logging: bool = True
) -> Optional[str]:
    """
    Generate a complete CSV report for a school.

    Args:
        school_name: School name (used for directory lookup)
        cache_base_dir: Base cache directory
        output_base_dir: Base output directory
        nmdp_db_path: Path to GITG database
        aliases_path: Path to school aliases
        config_path: Path to config file
        enable_logging: Whether to set up file logging

    Returns:
        Path to generated CSV file, or None if error
    """
    # Set up logging for this school
    if enable_logging:
        setup_logger(school_name=school_name)

    logger = get_logger()
    log_section(f"Generating Report for {school_name}")

    # Normalize school name for directory
    school_dir_name = school_name.lower().replace(" ", "_").replace("-", "_")
    coaches_dir = os.path.join(cache_base_dir, school_dir_name, "coaches")

    logger.info(f"Cache directory: {coaches_dir}")

    if not os.path.exists(coaches_dir):
        logger.error(f"Coaches directory not found: {coaches_dir}")
        return None

    # Load config
    config = {}
    if os.path.exists(config_path):
        config = load_json_file(config_path)
        logger.info(f"Loaded config from {config_path}")

    # Run cross-reference
    logger.info("Running cross-reference against NMDP database...")
    results = cross_reference_all_coaches(coaches_dir, nmdp_db_path, aliases_path, config_path)

    if not results:
        logger.warning(f"No coach data found in {coaches_dir}")
        return None

    logger.info(f"Cross-reference complete: {len(results)} coaches processed")

    # Generate output filenames
    date_str = datetime.now().strftime("%Y-%m-%d")
    csv_filename = f"{school_dir_name}_{date_str}.csv"
    csv_path = os.path.join(output_base_dir, csv_filename)

    # Generate CSV
    logger.info(f"Generating CSV report: {csv_path}")
    result_path = generate_csv_report(results, csv_path, config)

    # Generate Excel if available
    excel_path = None
    if EXCEL_AVAILABLE:
        excel_filename = f"{school_dir_name}_{date_str}.xlsx"
        excel_path = os.path.join(output_base_dir, excel_filename)
        logger.info(f"Generating Excel report with hyperlinks: {excel_path}")
        generate_excel_report(results, excel_path, config)
    else:
        logger.warning("openpyxl not available - skipping Excel output")

    # Calculate and log statistics
    stats = generate_summary_stats(results)
    log_section("Report Summary")
    log_summary({
        "Total coaches": stats["total_coaches"],
        "Coaches with NMDP overlap": stats["coaches_with_overlap"],
        "Total overlap instances": stats["total_overlap_instances"],
        "Unique overlap schools": stats["unique_overlap_schools"],
        "Data quality - Verified": stats["data_quality"]["verified"],
        "Data quality - Partial": stats["data_quality"]["partial"],
        "Data quality - Unverified": stats["data_quality"]["unverified"]
    })

    if stats["overlap_schools_list"]:
        logger.info(f"Overlap schools: {', '.join(stats['overlap_schools_list'])}")

    logger.info(f"CSV report saved to: {result_path}")
    if excel_path:
        logger.info(f"Excel report saved to: {excel_path}")

    if enable_logging:
        log_file = get_log_file_path()
        if log_file:
            logger.info(f"Log file saved to: {log_file}")

    # Also print to console for user feedback
    print(f"CSV report generated: {result_path}")
    if excel_path:
        print(f"Excel report generated: {excel_path} (with clickable hyperlinks)")
    print(f"Total coaches: {len(results)}")
    print(f"Coaches with NMDP overlap: {sum(1 for r in results if r.get('has_overlap'))}")

    return result_path, excel_path if excel_path else result_path


def generate_summary_stats(results: List[Dict]) -> Dict:
    """
    Generate summary statistics from results.

    Args:
        results: List of cross-reference result dictionaries

    Returns:
        Dictionary with summary statistics
    """
    total_coaches = len(results)
    coaches_with_overlap = sum(1 for r in results if r.get("has_overlap"))
    total_overlaps = sum(r.get("overlap_count", 0) for r in results)

    verified = sum(1 for r in results if determine_data_quality(r.get("career_history", []), r.get("research_status", "")) == "VERIFIED")
    partial = sum(1 for r in results if determine_data_quality(r.get("career_history", []), r.get("research_status", "")) == "PARTIAL")
    unverified = sum(1 for r in results if determine_data_quality(r.get("career_history", []), r.get("research_status", "")) == "UNVERIFIED")

    # Unique schools with overlaps
    overlap_schools = set()
    for r in results:
        for overlap in r.get("overlaps", []):
            overlap_schools.add(overlap.get("school"))

    return {
        "total_coaches": total_coaches,
        "coaches_with_overlap": coaches_with_overlap,
        "coaches_without_overlap": total_coaches - coaches_with_overlap,
        "total_overlap_instances": total_overlaps,
        "unique_overlap_schools": len(overlap_schools),
        "overlap_schools_list": sorted(overlap_schools),
        "data_quality": {
            "verified": verified,
            "partial": partial,
            "unverified": unverified
        },
        "overlap_percentage": (coaches_with_overlap / total_coaches * 100) if total_coaches > 0 else 0
    }


if __name__ == "__main__":
    """CLI usage: python generate_csv.py <school_name> [cache_dir] [output_dir]"""

    if len(sys.argv) < 2:
        print("Usage: python generate_csv.py <school_name> [cache_dir] [output_dir]")
        print("  school_name: Name of school (used to find cache directory)")
        print("  cache_dir: Base cache directory (default: ../cache)")
        print("  output_dir: Output directory (default: ../output)")
        sys.exit(1)

    school_name = sys.argv[1]

    # Default paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)

    cache_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(base_dir, "cache")
    output_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.join(base_dir, "output")

    nmdp_path = os.path.join(base_dir, "data", "gitg_school_years.json")
    aliases_path = os.path.join(base_dir, "data", "school_aliases.json")
    config_path = os.path.join(base_dir, "config.json")

    result = generate_report_for_school(
        school_name,
        cache_dir,
        output_dir,
        nmdp_path,
        aliases_path,
        config_path
    )

    if result:
        if isinstance(result, tuple):
            csv_path, excel_path = result
            print(f"\nCSV report saved to: {csv_path}")
            if excel_path and excel_path != csv_path:
                print(f"Excel report saved to: {excel_path}")
        else:
            print(f"\nReport saved to: {result}")
    else:
        print("\nFailed to generate report")
        sys.exit(1)
