"""
import_territory_data.py - Import territory data from NMDP Excel file

Reads the CA and TX Coverage Map Excel file and generates territory_mapping.json
"""

import json
import os
import openpyxl
from collections import defaultdict


def import_territories(excel_path: str, output_path: str):
    """
    Import territory data from the NMDP coverage Excel file.

    Creates a mapping structure that supports:
    - State-level territory assignment (for most states)
    - County-level territory assignment (for CA and TX)
    """
    wb = openpyxl.load_workbook(excel_path)

    territories = {
        "_source": excel_path,
        "_updated": "2026-01-31",
        "state_territories": {},
        "county_territories": {
            "California": {},
            "Texas": {}
        },
        "territory_details": {}
    }

    # === Extract West Region state mappings ===
    ws = wb['West Region FY25']
    for row in ws.iter_rows(min_row=9, max_row=30, values_only=True):
        state = row[1]
        territory = row[2]
        am = row[3]  # Account Manager

        if state and territory and not state.startswith('AANHPI') and not state.startswith('Gulf') and not state.startswith('Moving') and not state.startswith('North Texas') and not state.startswith('Northern CA') and not state.startswith('Southern CA'):
            # Clean up state name
            state_clean = state.strip()
            territory_clean = territory.strip()

            territories["state_territories"][state_clean] = territory_clean

            # Track territory details
            if territory_clean not in territories["territory_details"]:
                territories["territory_details"][territory_clean] = {
                    "states": [],
                    "account_managers": set()
                }
            if state_clean not in territories["territory_details"][territory_clean]["states"]:
                territories["territory_details"][territory_clean]["states"].append(state_clean)
            if am:
                territories["territory_details"][territory_clean]["account_managers"].add(am)

    # === Extract California county mappings ===
    ws = wb['California Coverage by County']
    for row in ws.iter_rows(min_row=5, values_only=True):
        territory = row[2]  # Territory column
        county = row[3]     # County column
        am = row[4]         # Account Manager

        if county and territory:
            # Clean county name (remove ", CA")
            county_clean = county.replace(", CA", "").strip()
            territory_clean = territory.strip()

            territories["county_territories"]["California"][county_clean] = territory_clean

            # Track territory details
            if territory_clean not in territories["territory_details"]:
                territories["territory_details"][territory_clean] = {
                    "states": ["California"],
                    "account_managers": set()
                }
            if am:
                territories["territory_details"][territory_clean]["account_managers"].add(am)

    # === Extract Texas county mappings ===
    # Texas uses AM names as implicit territories
    ws = wb['Texas Coverage by County']

    # First, map AMs to territory names
    texas_am_territory = {
        "Ryan Dixon": "North Texas",
        "Brian Allison": "East Texas / Houston",
        "Leticia Mondragon": "South Texas / San Antonio",
        "Miranda Robinson": "Gulf Coast",
        "Eric Bolton": "East Texas / Houston"
    }

    for row in ws.iter_rows(min_row=8, values_only=True):
        county = row[0]
        am = row[1]

        if county and am:
            county_clean = county.replace(", TX", "").strip()
            territory = texas_am_territory.get(am, "Texas - Other")

            territories["county_territories"]["Texas"][county_clean] = territory

            # Track territory details
            if territory not in territories["territory_details"]:
                territories["territory_details"][territory] = {
                    "states": ["Texas"],
                    "account_managers": set()
                }
            territories["territory_details"][territory]["account_managers"].add(am)

    # Convert sets to lists for JSON serialization
    for territory_name, details in territories["territory_details"].items():
        details["account_managers"] = sorted(list(details["account_managers"]))

    # Add additional manual mappings from the territoryinformation file
    # These override/supplement the Excel data
    manual_overrides = {
        "Alaska": "Sacramento Valley & North Coast",
        "Hawaii": "Southern CA West",  # Note: Excel says "Hawaii" but text file says Southern CA West
        "Wyoming": "Northwest",
        "New Mexico": "Southwest"
    }
    for state, territory in manual_overrides.items():
        territories["state_territories"][state] = territory

    # Save to JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(territories, f, indent=2, ensure_ascii=False)

    print(f"Territory mapping saved to: {output_path}")
    print(f"\nSummary:")
    print(f"  State-level mappings: {len(territories['state_territories'])}")
    print(f"  California county mappings: {len(territories['county_territories']['California'])}")
    print(f"  Texas county mappings: {len(territories['county_territories']['Texas'])}")
    print(f"  Total territories: {len(territories['territory_details'])}")

    print(f"\nTerritories:")
    for t_name in sorted(territories['territory_details'].keys()):
        details = territories['territory_details'][t_name]
        print(f"  - {t_name}")
        if details['states']:
            print(f"      States: {', '.join(details['states'])}")
        if details['account_managers']:
            print(f"      AMs: {', '.join(details['account_managers'])}")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)

    excel_path = os.path.join(base_dir, "data", "CA and TX Coverage Map for Lead Distribution (1).xlsx")
    output_path = os.path.join(base_dir, "data", "territory_mapping.json")

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return

    import_territories(excel_path, output_path)


if __name__ == "__main__":
    main()
