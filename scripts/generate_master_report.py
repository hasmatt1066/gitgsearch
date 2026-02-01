"""
generate_master_report.py - Master Excel report with territory tabs

Generates a consolidated Excel workbook with:
- Master tab: All results from all school searches
- Territory tabs: Filtered views by NMDP territory

Each row includes state/county and territory columns.
"""

import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set

from cross_reference import (
    cross_reference_all_coaches,
    format_overlaps_summary,
    load_json_file
)
from generate_csv import (
    get_career_entries_with_urls,
    determine_data_quality,
    format_career_history,
    EXCEL_AVAILABLE
)

if EXCEL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def load_school_locations(locations_path: str) -> Dict:
    """Load school -> state/county mapping."""
    if os.path.exists(locations_path):
        return load_json_file(locations_path)
    return {}


def load_territory_mapping(territory_path: str) -> Dict:
    """Load territory mapping (state and county level)."""
    if os.path.exists(territory_path):
        return load_json_file(territory_path)
    return {}


def load_school_aliases(aliases_path: str) -> Dict[str, List[str]]:
    """Load school aliases for name matching."""
    if os.path.exists(aliases_path):
        return load_json_file(aliases_path)
    return {}


def normalize_school_name_for_lookup(school_name: str, aliases: Dict[str, List[str]]) -> str:
    """
    Try to find the canonical NMDP name for a school.

    Args:
        school_name: The school name to look up
        aliases: The aliases dictionary

    Returns:
        Canonical name if found, otherwise original name uppercased
    """
    # Check if it's already a canonical name
    upper_name = school_name.upper()
    if upper_name in aliases:
        return upper_name

    # Check if it matches any alias
    for canonical, alias_list in aliases.items():
        if canonical.startswith("_"):
            continue
        for alias in alias_list:
            if alias.upper() == upper_name or alias.upper() in upper_name.upper():
                return canonical

    return upper_name


def get_school_location(school_name: str, locations: Dict, aliases: Dict) -> Tuple[str, str]:
    """
    Get state and county for a school.

    Args:
        school_name: School name (may be canonical or alias)
        locations: Location mapping dictionary
        aliases: School aliases dictionary

    Returns:
        Tuple of (state, county) or ("Unknown", "Unknown") if not found
    """
    # Try direct lookup first
    if school_name in locations:
        loc = locations[school_name]
        return loc.get("state", "Unknown"), loc.get("county", "Unknown")

    # Try uppercase
    upper_name = school_name.upper()
    if upper_name in locations:
        loc = locations[upper_name]
        return loc.get("state", "Unknown"), loc.get("county", "Unknown")

    # Try canonical name lookup
    canonical = normalize_school_name_for_lookup(school_name, aliases)
    if canonical in locations:
        loc = locations[canonical]
        return loc.get("state", "Unknown"), loc.get("county", "Unknown")

    return "Unknown", "Unknown"


def get_territory_for_location(state: str, county: str, territories: Dict) -> str:
    """
    Get NMDP territory for a state/county combination.

    For California and Texas, uses county-level mapping.
    For other states, uses state-level mapping.

    Args:
        state: State name
        county: County name
        territories: Territory mapping dictionary

    Returns:
        Territory name or "Unknown" if not found
    """
    if state == "Unknown":
        return "Unknown"

    # Check for county-level mapping (California and Texas)
    county_territories = territories.get("county_territories", {})

    if state == "California" and "California" in county_territories:
        ca_counties = county_territories["California"]
        # Try exact match first
        if county in ca_counties:
            return ca_counties[county]
        # Try without "County" suffix
        county_base = county.replace(" County", "").strip()
        if county_base in ca_counties:
            return ca_counties[county_base]
        # Try with "County" suffix
        if f"{county_base} County" in ca_counties:
            return ca_counties[f"{county_base} County"]

    if state == "Texas" and "Texas" in county_territories:
        tx_counties = county_territories["Texas"]
        # Try exact match first
        if county in tx_counties:
            return tx_counties[county]
        # Try without "County" suffix
        county_base = county.replace(" County", "").strip()
        if county_base in tx_counties:
            return tx_counties[county_base]
        # Try with "County" suffix
        if f"{county_base} County" in tx_counties:
            return tx_counties[f"{county_base} County"]

    # Fall back to state-level mapping
    state_territories = territories.get("state_territories", {})
    if state in state_territories:
        return state_territories[state]

    return "Unknown"


def get_all_cached_schools(cache_base_dir: str) -> List[str]:
    """
    Get list of all schools with cached data.

    Args:
        cache_base_dir: Base cache directory

    Returns:
        List of school directory names
    """
    schools = []
    if os.path.exists(cache_base_dir):
        for item in os.listdir(cache_base_dir):
            item_path = os.path.join(cache_base_dir, item)
            if os.path.isdir(item_path):
                coaches_dir = os.path.join(item_path, "coaches")
                if os.path.exists(coaches_dir):
                    schools.append(item)
    return sorted(schools)


def aggregate_all_results(
    cache_base_dir: str,
    nmdp_db_path: str,
    aliases_path: str,
    config_path: str,
    locations_path: str,
    territory_path: str
) -> List[Dict]:
    """
    Aggregate cross-reference results from all cached schools.

    Returns list of result dictionaries, each augmented with:
    - searched_school: The school that was searched
    - state: State of the searched school
    - county: County of the searched school
    - territory: NMDP territory
    """
    # Load mappings
    locations = load_school_locations(locations_path)
    territories = load_territory_mapping(territory_path)
    aliases = load_school_aliases(aliases_path)
    config = load_json_file(config_path) if os.path.exists(config_path) else {}

    all_results = []
    schools = get_all_cached_schools(cache_base_dir)

    print(f"Found {len(schools)} schools with cached data")

    for school_dir in schools:
        coaches_dir = os.path.join(cache_base_dir, school_dir, "coaches")

        # Cross-reference this school's coaches
        results = cross_reference_all_coaches(
            coaches_dir, nmdp_db_path, aliases_path, config_path
        )

        if not results:
            print(f"  {school_dir}: No coach data")
            continue

        # Get the current school name from first coach (they should all be the same)
        current_school = results[0].get("current_school", school_dir.replace("_", " ").title())

        # Look up location and territory
        state, county = get_school_location(current_school, locations, aliases)
        territory = get_territory_for_location(state, county, territories)

        print(f"  {school_dir}: {len(results)} coaches, {state}, {territory}")

        # Augment each result
        for result in results:
            result["searched_school"] = current_school
            result["state"] = state
            result["county"] = county
            result["territory"] = territory
            all_results.append(result)

    return all_results


def write_sheet_data(
    ws,
    results: List[Dict],
    headers: List[str],
    max_career_entries: int,
    year_start: int,
    styles: Dict
):
    """
    Write data rows to a worksheet.

    Args:
        ws: Worksheet to write to
        results: List of result dictionaries
        headers: Column headers
        max_career_entries: Number of career columns
        year_start: Earliest year for career filtering
        styles: Dictionary of styles (header_font, header_fill, etc.)
    """
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = styles["thin_border"]

    # Write data rows
    for row_idx, result in enumerate(results, 2):
        career_history = result.get("career_history", [])
        career_entries = get_career_entries_with_urls(career_history, year_start)
        has_overlap = result.get("has_overlap", False)

        col = 1

        # School searched
        ws.cell(row=row_idx, column=col, value=result.get("searched_school", "Unknown"))
        col += 1

        # State
        ws.cell(row=row_idx, column=col, value=result.get("state", "Unknown"))
        col += 1

        # County
        ws.cell(row=row_idx, column=col, value=result.get("county", "Unknown"))
        col += 1

        # Territory
        ws.cell(row=row_idx, column=col, value=result.get("territory", "Unknown"))
        col += 1

        # Coach info
        ws.cell(row=row_idx, column=col, value=result.get("coach_name", "Unknown"))
        col += 1

        ws.cell(row=row_idx, column=col, value=result.get("current_position", "Unknown"))
        col += 1

        # Career columns with hyperlinks
        for i in range(max_career_entries):
            if i < len(career_entries):
                display_text, source_url = career_entries[i]
                cell = ws.cell(row=row_idx, column=col, value=display_text)
                if source_url:
                    cell.hyperlink = source_url
                    cell.font = styles["link_font"]
            else:
                ws.cell(row=row_idx, column=col, value="")
            col += 1

        # Overlap columns
        ws.cell(row=row_idx, column=col, value="YES" if has_overlap else "NO")
        col += 1

        ws.cell(row=row_idx, column=col, value=format_overlaps_summary(result.get("overlaps", [])))
        col += 1

        ws.cell(row=row_idx, column=col, value=determine_data_quality(career_history, result.get("research_status", "")))
        col += 1

        # Apply borders and highlighting
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = styles["thin_border"]
            if has_overlap:
                cell.fill = styles["overlap_fill"]

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(results) + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def generate_master_report(
    cache_base_dir: str,
    output_path: str,
    nmdp_db_path: str,
    aliases_path: str,
    config_path: str,
    locations_path: str,
    territory_path: str
) -> Optional[str]:
    """
    Generate master Excel report with all schools and territory tabs.

    Args:
        cache_base_dir: Base cache directory
        output_path: Path to write Excel file
        nmdp_db_path: Path to GITG database
        aliases_path: Path to school aliases
        config_path: Path to config file
        locations_path: Path to school locations file
        territory_path: Path to territory mapping file

    Returns:
        Path to generated Excel file, or None if error
    """
    if not EXCEL_AVAILABLE:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        return None

    print("Aggregating all cached school data...")
    all_results = aggregate_all_results(
        cache_base_dir, nmdp_db_path, aliases_path, config_path,
        locations_path, territory_path
    )

    if not all_results:
        print("No results found in cache")
        return None

    print(f"\nTotal coaches across all schools: {len(all_results)}")

    # Load config for year range
    config = load_json_file(config_path) if os.path.exists(config_path) else {}
    year_start = config.get("year_range", {}).get("start", 2020)

    # Calculate max career entries
    max_career_entries = 0
    for result in all_results:
        entries = get_career_entries_with_urls(result.get("career_history", []), year_start)
        max_career_entries = max(max_career_entries, len(entries))
    max_career_entries = max(max_career_entries, 3)

    # Build headers
    headers = [
        "School Searched",
        "State",
        "County",
        "Territory",
        "Coach Name",
        "Current Position"
    ]
    for i in range(max_career_entries):
        headers.append(f"Career {i+1}")
    headers.extend(["NMDP Overlap", "Overlap Details", "Data Quality"])

    # Define styles
    styles = {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "overlap_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "link_font": Font(color="0563C1", underline="single"),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "thin_border": Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    # Create workbook
    wb = Workbook()

    # Master tab (all results)
    ws_master = wb.active
    ws_master.title = "Master - All Results"
    write_sheet_data(ws_master, all_results, headers, max_career_entries, year_start, styles)

    # Get unique territories
    territories = set()
    for result in all_results:
        territory = result.get("territory", "Unknown")
        if territory:
            territories.add(territory)

    # Create tab for each territory
    for territory in sorted(territories):
        # Filter results for this territory
        territory_results = [r for r in all_results if r.get("territory") == territory]

        if not territory_results:
            continue

        # Create sheet with sanitized name (Excel limits to 31 chars)
        sheet_name = territory[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=sheet_name)

        write_sheet_data(ws, territory_results, headers, max_career_entries, year_start, styles)

        print(f"  {territory}: {len(territory_results)} coaches")

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save workbook
    wb.save(output_path)

    # Print summary
    print(f"\n{'='*50}")
    print("MASTER REPORT SUMMARY")
    print(f"{'='*50}")
    print(f"Total schools: {len(get_all_cached_schools(cache_base_dir))}")
    print(f"Total coaches: {len(all_results)}")
    print(f"Coaches with NMDP overlap: {sum(1 for r in all_results if r.get('has_overlap'))}")
    print(f"Territories: {len(territories)}")
    print(f"\nReport saved to: {output_path}")

    return output_path


def main():
    """CLI entry point."""
    # Default paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)

    cache_dir = os.path.join(base_dir, "cache")
    output_dir = os.path.join(base_dir, "output")

    nmdp_path = os.path.join(base_dir, "data", "gitg_school_years.json")
    aliases_path = os.path.join(base_dir, "data", "school_aliases.json")
    config_path = os.path.join(base_dir, "config.json")
    locations_path = os.path.join(base_dir, "data", "school_locations.json")
    territory_path = os.path.join(base_dir, "data", "territory_mapping.json")

    # Generate output filename
    date_str = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(output_dir, f"master_report_{date_str}.xlsx")

    result = generate_master_report(
        cache_dir,
        output_path,
        nmdp_path,
        aliases_path,
        config_path,
        locations_path,
        territory_path
    )

    if result:
        print(f"\nSuccess! Master report: {result}")
    else:
        print("\nFailed to generate master report")
        sys.exit(1)


if __name__ == "__main__":
    main()
